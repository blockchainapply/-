# -*- coding: utf-8 -*-

import requests
import xlwt
from bs4 import BeautifulSoup
import time
import datetime
import pandas as pd
import re
import xlsxwriter
from openpyxl import load_workbook
from nltk.stem import PorterStemmer
from nltk.tokenize import RegexpTokenizer
from nltk.corpus import stopwords
from string import digits
import matplotlib.pyplot as plt
from sklearn.decomposition import PCA
from sklearn import svm
from sklearn import metrics
import os
import gensim
import nltk
from gensim.models import word2vec
import xlrd
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
import re
import time
start=time.clock()
url = 'http://www.chinadaily.com.cn/business/money'

headers = {
        'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding':'gzip, deflate',
        'Accept-Language':'zh-CN,zh;q=0.9',
        'Cache-Control':'max-age=0',
        'Connection':'keep-alive',
        'Host':'www.chinadaily.com.cn',
        'Upgrade-Insecure-Requests':'1',
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36'
        }#正常的浏览器在给它发送信息
r = requests.get(url,headers=headers) #获取页面信息，防止反爬虫
soup = BeautifulSoup(r.text,'lxml')
infos = soup.find_all('span',attrs={'class':'tw3_01_2_p'})[:1]
for i in range(len(infos)):
    print('正在爬取第{}个...'.format(i+1))
    url = 'http:' + infos[i].find('a')['href']
    r = requests.get(url,headers=headers)#获得界面
    soup = BeautifulSoup(r.text,'lxml')
    text = soup.find('div',attrs={'id':'Content'}).text.replace('\n','')#获得这个页面结构中的id和content
    title = soup.find('h1').text.strip()
    #file_name = title +' '+str(datetime.datetime.now())
    with open(str(i+1)+'.txt','a',encoding='utf-8') as f:
        f.write(title)
        f.write(text)
    time.sleep(2)

with open('1.txt','r',encoding="utf-8") as f:
    content=f.read()
    contents=re.split('\s|\.|\"|,',content)#txt拆分，以 空格 。" '
    final=list(filter(None,contents))
f.close()
workbook = xlsxwriter.Workbook('C:/Users/Administrator/Desktop/2018.xlsx')#创建excel
worksheet = workbook.add_worksheet()#创建sheet
row = 0
col = 1
all_cols=len(final)//300
for i in range(all_cols+1):
    worksheet.write(i+1,0, " ".join(final[i*300:(i+1)*300]))#以""为分隔符，将字符串连起来从0-300
    col+=1
worksheet.write(0,0,'comment')
workbook.close()


filename = '2018.xlsx'
model = gensim.models.KeyedVectors.load_word2vec_format('C:/Users/Administrator/Desktop/'
                                                       'GoogleNews-vectors-negative300-SLIM.bin.gz', binary=True)

fileVecs = []
file = pd.read_excel(filename, header=0)
com_num = len(file) # 记录一共有多少文段
# 下明tokenizer，stop_words， remove_digits 是为后面词块化和去除停用词、非字母字符做准备
tokenizer = RegexpTokenizer(r'\w+')
#stemmer = PorterStemmer()
stop_words = set(stopwords.words('english'))
remove_digits = str.maketrans('', '', digits)
sents= []  # 定义一个列表，添加处理后的每个文段
# 对每段评论做处理，目的是针对每个文段，最后都处理为一个去除 停用词、数字、标点 之后的单词串
for i in range(com_num):
    sent = file.loc[i, 'content']  # 1.取一行评论
    sent = str(sent)#2.把这行评论变成string
    input_str1 = sent.lower() #3.把字母都转化为小写
    input_str2 = input_str1.translate(remove_digits) #4.去数字
    raw_tokens = tokenizer.tokenize(input_str2) # 5.句子词块化
    stemmed_tokens_without_stopword = filter(lambda x: x not in stop_words, raw_tokens) #6.去停用词
    stemmed = list(stemmed_tokens_without_stopword) #7.把处理后的文字变成list形式
    sents.append(stemmed) # 8.列表，append处理后的每个文段

comment_data = pd.Series(sents)  # list转pandas，内容没变，主要是转转形式为了后面好处理
file['com_tok'] = comment_data  # file在最后一列后面再添加一列名为com_tok，内容就是每个文段处理后的样子
vec_list = []
word_count = 0
vec2 = 0
c = 0

for line in comment_data:  # 对于每个处理后的文段
    c += 1
    for word in line:  # 找到单条评论里每个单词
        if word in model:  # 如果这个单词在model里
            word_count += 1  # 针对这个评论的单词计数器+1
            vec1 = model[word]  # 这个单词在model中找到自己对应的词向量 例：[12，45，667，980，...，278，10]，每个单词在model中对应的向量等长（维度相同）
            vec2 = vec1+vec2  # 评论里每个单词对应的向量相加
    if word_count == 0:  # 异常情况处理
        file = file.drop([c-1])
        continue
    else:
        avg_vec_per_sent=vec2/word_count  # 把加和后的向量除以单词数量（即取平均数），而后得到代表这段文段的向量值
        vec_list.append(avg_vec_per_sent)  # vec_list记录每个文段的向量值
    word_count = 0  # 单词计数器清零，准备迎接下一个评论
    vec2 = 0
print(vec_list[:c])

book = load_workbook(filename=r"C:/Users/Administrator/Desktop/test.xlsx")
sheet = book.get_sheet_by_name("Sheet1")
  #用于存储数据的数组
num = 1
while 1:
    cell = sheet.cell(row=num, column=1).value
    if cell:
        num = num +1
    else:
       print(num-1)
       break

def cosVector(x,y):
    if(len(x)!=len(y)):
        print('error input,x and y is not in the same space')
        return;
    result1=0.0;
    result2=0.0;
    result3=0.0;
    for i in range(len(x)):

        result1+=x[i]*y[i]   #sum(X*Y)
        result2+=x[i]**2     #sum(X*X)
        result3+=y[i]**2     #sum(Y*Y)
    return (float(result1)/((float(result2)*float(result3))**0.5))#结果显示
data= []
row_num = 1#行
column_num = 1
for i in range(len(vec_list[:c])):
    for data_num in range(num-1):
        while column_num <= 300 :
        #将表中300列数据写入data数组中
            data.append(sheet.cell(row=row_num, column=column_num).value)
            column_num = column_num+1
        label = [float(d) for d in data]
        print('\n')
        a=float(cosVector(label,vec_list[i]))#分别计算其相似度作对比
        print(a)
        if a<0.9:
            row_num=row_num+1
            column_num=1
            data=[]
            label=[]
        else:
            exit('该文章相似度较高')
    row_num = 1  # 行
    column_num = 1

workbook2=xlrd.open_workbook('test.xlsx')
sheet2=workbook2.sheet_by_index(0)
row=sheet2.nrows

workbook=load_workbook('test.xlsx')
sheet=workbook['Sheet1']
for m in range(len(vec_list[:c])):
    for i in range(0,300):
        sheet.cell(num,i+1,str(vec_list[m][i]))
    num=num+1
    i=1
workbook.save('test.xlsx')



sum_vec_list = 0
avg_vet_list = 0
str_avg_vec_per_sent=0

for i in range(len(vec_list)):
    sum_vec_list += vec_list[i]
avg_vec_per_sent = sum_vec_list / len(vec_list)
str_avg_vec_per_sent= str(avg_vec_per_sent)

print(str_avg_vec_per_sent)

with open("C:/Users/Administrator/Desktop/test.txt", "w", encoding='utf-8') as f:
    f.write(str_avg_vec_per_sent)
end=time.clock()
print("final time used ",end-start)


